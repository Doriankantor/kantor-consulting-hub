#!/usr/bin/env python3
"""Simple, human-readable version of the poolable-outcomes table.
Outputs a clean colour-coded Excel + a plain CSV for quick visual scanning."""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_metafor_table import ROWS  # (study, subgroup, outcome, measure, eff, lo, hi, adj, p0, conv, flag)

HERE = os.path.dirname(os.path.abspath(__file__))

CARE = {"HbA1c_control", "HbA1c_monitoring", "Retinopathy_screening"}   # higher = better care
EVENT = {"All-cause_mortality", "CV_disease"}                            # higher = worse
OG_ORDER = ["HbA1c_control", "HbA1c_monitoring", "Retinopathy_screening",
            "All-cause_mortality", "CV_disease"]
OG_LABEL = {
    "HbA1c_control": "HbA1c control (achieving glycaemic target)",
    "HbA1c_monitoring": "HbA1c monitoring (received testing)",
    "Retinopathy_screening": "Retinopathy / eye screening (received)",
    "All-cause_mortality": "All-cause mortality  [post-hoc]",
    "CV_disease": "Cardiovascular disease / events",
}


def favors(og, eff, lo, hi, flag):
    if lo == "" or hi == "":
        return "Inconclusive (no CI)"
    lo, hi = float(lo), float(hi)
    sig = (lo > 1 and hi > 1) or (lo < 1 and hi < 1)   # CI excludes 1
    if not sig:
        return "Inconclusive"
    low_is_worse = og in CARE
    res = ("Worse in migrants" if eff < 1 else "Better in migrants") if low_is_worse \
        else ("Better in migrants" if eff < 1 else "Worse in migrants")
    if "INVERSE" in flag:                              # outcome measures the complement
        res = {"Worse in migrants": "Better in migrants",
               "Better in migrants": "Worse in migrants"}.get(res, res)
    return res


def ci_str(lo, hi):
    return f"{lo}–{hi}" if lo != "" and hi != "" else "not reported"


# ---- assemble simple rows ----
simple = []
for st, sub, og, meas, eff, lo, hi, adj, p0, conv, flag in ROWS:
    simple.append({
        "outcome": og,
        "study": st.replace("_", " "),
        "group": "" if sub == "overall" else sub.replace("_", " "),
        "measure": meas,
        "effect": eff,
        "ci": ci_str(lo, hi),
        "favors": favors(og, eff, lo, hi, flag),
    })

# ---- plain CSV ----
csv_path = os.path.join(HERE, "outcomes_table_simple.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Outcome", "Study", "Migrant group", "Measure", "Effect", "95% CI", "Favours"])
    for og in OG_ORDER:
        for r in [x for x in simple if x["outcome"] == og]:
            w.writerow([OG_LABEL[og], r["study"], r["group"], r["measure"],
                        r["effect"], r["ci"], r["favors"]])

# ---- pretty Excel ----
wb = Workbook(); ws = wb.active; ws.title = "Poolable outcomes"
NAVY = "1F4E79"; LBLUE = "E8F1FB"
FILL_WORSE = PatternFill("solid", fgColor="F4CCCC")
FILL_BETTER = PatternFill("solid", fgColor="D9EAD3")
FILL_NS = PatternFill("solid", fgColor="EFEFEF")
hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
grp_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Study", "Migrant group", "Measure", "Effect", "95% CI", "Favours"]
ws.append([""])  # title row placeholder
ws["A1"] = "Poolable outcomes (≥3 studies) — effect of migrant status, one row per study/subgroup"
ws["A1"].font = Font(bold=True, size=13, color=NAVY)
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(2, c); cell.font = hdr_font; cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border

r = 3
for og in OG_ORDER:
    rows = [x for x in simple if x["outcome"] == og]
    ws.cell(r, 1, OG_LABEL[og]).font = grp_font
    for c in range(1, len(headers) + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor="2E75B6")
        ws.cell(r, c).border = border
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    r += 1
    for x in rows:
        vals = [x["study"], x["group"], x["measure"], x["effect"], x["ci"], x["favors"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.border = border
            cell.alignment = Alignment(horizontal="center" if c >= 3 else "left", vertical="center")
        fill = (FILL_WORSE if x["favors"].startswith("Worse")
                else FILL_BETTER if x["favors"].startswith("Better") else FILL_NS)
        ws.cell(r, 6).fill = fill
        if r % 2 == 0:
            for c in (1, 2, 3, 4, 5):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=LBLUE)
        r += 1

widths = [22, 16, 9, 9, 16, 20]
for i, wdt in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = wdt
ws.freeze_panes = "A3"
xlsx_path = os.path.join(HERE, "outcomes_table_simple.xlsx")
wb.save(xlsx_path)

print("Wrote:\n ", csv_path, "\n ", xlsx_path, f"\n  ({len(simple)} rows)")
