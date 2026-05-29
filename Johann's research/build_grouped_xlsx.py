#!/usr/bin/env python3
"""Export outcomes grouped by clinical concept to an Excel sheet."""
import os
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "diabetes_migrants.db")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "diabetes_outcomes_grouped.xlsx")

# clinical-flow order of categories
CAT_ORDER = ["Process-of-Care", "Glycemic Control", "Microvascular",
             "Macrovascular", "Mortality", "Acute"]


def outcome_group(cat, name):
    n = (name or "").lower()
    if cat == "Acute":
        return "Acute metabolic complications"
    if cat == "Glycemic Control":
        return "Glycosuria" if "glycosuria" in n else "HbA1c glycaemic control (target/level)"
    if cat == "Mortality":
        if "cardiovascular" in n:
            return "Cardiovascular mortality"
        if "diabetes-related" in n or "diabetes-specific" in n or "dm-specific" in n:
            return "Diabetes-specific mortality"
        return "All-cause mortality"
    if cat == "Microvascular":
        if any(k in n for k in ("retinopathy", "retinal", "vitrectomy", "ophthalmolog", "blindness")):
            return "Retinopathy (disease & treatment)"
        if any(k in n for k in ("nephropathy", "renal", "acr", "gfr")):
            return "Nephropathy / renal"
        if any(k in n for k in ("podiatric", "amputation", "foot")):
            return "Diabetic foot / amputation"
        return "Other microvascular"
    if cat == "Macrovascular":
        if "stroke" in n:
            return "Stroke"
        if "chd" in n or "coronary" in n:
            return "Coronary heart disease (CHD)"
        return "Cardiovascular disease / composite CV events"
    if cat == "Process-of-Care":
        if any(k in n for k in ("readmission", "hospitaliz", "hospital stay", "length of hospital")):
            return "Hospital utilization (admission/readmission/LOS)"
        if "gp visit" in n:
            return "Primary-care visits (GP)"
        if "statin" in n or "acei" in n or "arb" in n or "insulin prescription" in n:
            return "Guideline medication prescribing"
        if "hba1c" in n and any(k in n for k in ("test", "measurement", "monitoring", "no hba1c value")):
            return "HbA1c testing/monitoring"
        if "ldl" in n and ("monitor" in n or "testing" in n):
            return "Lipid testing/monitoring"
        if "ldl" in n and ("target" in n or "≤2.0" in n or "<2.0" in n):
            return "LDL / lipid target attainment"
        if "sbp" in n and "monitor" in n:
            return "Blood-pressure monitoring"
        if "sbp" in n and "target" in n:
            return "Blood-pressure target attainment"
        if "kidney" in n or "egfr" in n or "acr" in n:
            return "Kidney-function monitoring (eGFR/ACR)"
        if "foot" in n:
            return "Foot examination"
        if any(k in n for k in ("ophthalmolog", "retinopathy screening", "eye examination", "retinopathy screening exam")):
            return "Retinopathy / eye screening"
        return "Other process-of-care"
    return "Uncategorized"


COLS = [
    ("Category", "outcome_category_norm", 17),
    ("Outcome group (clinical concept)", "_group", 34),
    ("Outcome ID", "outcome_id", 9),
    ("Outcome name", "outcome_name", 46),
    ("Comparison", "comparison_desc", 30),
    ("Measure", "measure_type_norm", 9),
    ("Migrants value", "migrants_value", 12),
    ("Comparator value", "comparator_value", 13),
    ("Effect estimate", "effect_estimate", 12),
    ("CI low", "ci_low", 8),
    ("CI high", "ci_high", 8),
    ("P-value", "p_value", 9),
    ("Significance", "significance", 14),
    ("Direction", "direction_norm", 18),
    ("Magnitude", "magnitude", 30),
    ("Supports hyp.", "supports_hypothesis", 13),
    ("Study ID", "study_id", 8),
    ("Author", "author", 30),
    ("Year", "year", 6),
    ("Destination country", "dest_country", 26),
    ("Health system", "health_system_type", 16),
    ("Study design", "study_design", 24),
    ("Quality", "quality_class", 9),
]
NUMERIC = {"migrants_value", "comparator_value", "effect_estimate", "ci_low", "ci_high"}

# light fill per category to make group blocks visually distinct
CAT_FILL = {
    "Process-of-Care": "E8F1FB", "Glycemic Control": "FDF3E0",
    "Microvascular": "EAF6F1", "Macrovascular": "FBEAE6",
    "Mortality": "EFEFEF", "Acute": "F3ECF7",
}


def main():
    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT o.*, s.author, s.year, s.dest_country, s.health_system_type,
               s.study_design, s.quality_class
        FROM outcomes o JOIN studies s ON s.study_id = o.study_id
    """).fetchall()
    conn.close()

    recs = []
    for r in rows:
        d = dict(r)
        d["_group"] = outcome_group(d["outcome_category_norm"], d["outcome_name"])
        recs.append(d)
    recs.sort(key=lambda d: (CAT_ORDER.index(d["outcome_category_norm"]),
                             d["_group"], d["study_id"], d["outcome_id"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Outcomes grouped"

    arial = "Arial"
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(name=arial, bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header
    for c, (label, _, width) in enumerate(COLS, 1):
        cell = ws.cell(1, c, label)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width

    # data rows
    for i, d in enumerate(recs):
        rownum = i + 2
        cat = d["outcome_category_norm"]
        fill = PatternFill("solid", fgColor=CAT_FILL.get(cat, "FFFFFF"))
        for c, (_, key, _) in enumerate(COLS, 1):
            val = d.get(key)
            cell = ws.cell(rownum, c, val if val is not None else "")
            cell.font = Font(name=arial, size=9)
            cell.border = border
            cell.fill = fill
            if key in NUMERIC and isinstance(val, (int, float)):
                cell.number_format = "0.###"
                cell.alignment = Alignment(horizontal="center")
            elif key in ("outcome_id", "study_id", "year", "p_value", "measure_type_norm",
                         "quality_class", "significance"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=key in
                                           ("outcome_name", "comparison_desc", "magnitude",
                                            "dest_country", "study_design"))
        # emphasize the two grouping columns
        for c in (1, 2):
            ws.cell(rownum, c).font = Font(name=arial, size=9, bold=True)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(recs) + 1}"
    ws.row_dimensions[1].height = 30

    # --- summary sheet ---
    ws2 = wb.create_sheet("Group summary")
    summ = {}
    for d in recs:
        k = (d["outcome_category_norm"], d["_group"])
        s = summ.setdefault(k, {"n": 0, "Worse in migrants": 0, "Better in migrants": 0,
                                "Inconclusive": 0, "Equal": 0, "sig": 0})
        s["n"] += 1
        if d["direction_norm"] in s:
            s[d["direction_norm"]] += 1
        if d["significance"] == "Significant":
            s["sig"] += 1
    heads = ["Category", "Outcome group", "Outcomes", "Worse", "Better",
             "Inconclusive", "Equal", "Significant"]
    widths2 = [17, 40, 9, 8, 8, 13, 7, 11]
    for c, (h, w) in enumerate(zip(heads, widths2), 1):
        cell = ws2.cell(1, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws2.column_dimensions[get_column_letter(c)].width = w
    ordered = sorted(summ, key=lambda k: (CAT_ORDER.index(k[0]), -summ[k]["n"], k[1]))
    for i, k in enumerate(ordered):
        s = summ[k]
        rownum = i + 2
        fill = PatternFill("solid", fgColor=CAT_FILL.get(k[0], "FFFFFF"))
        vals = [k[0], k[1], s["n"], s["Worse in migrants"], s["Better in migrants"],
                s["Inconclusive"], s["Equal"], s["sig"]]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(rownum, c, v)
            cell.font = Font(name=arial, size=9, bold=(c <= 2))
            cell.border = border
            cell.fill = fill
            if c >= 3:
                cell.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 28

    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Outcome rows: {len(recs)}")
    print(f"Clinical groups: {len(summ)}")
    for k in ordered:
        print(f"  {k[0]:17} | {k[1]:48} | n={summ[k]['n']}")


if __name__ == "__main__":
    main()
